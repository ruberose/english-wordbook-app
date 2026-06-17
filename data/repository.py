"""xlsx 파일 읽기/쓰기 담당 레포지토리.

UI에 의존하지 않는 순수 데이터 레이어 — 모바일/웹 재사용 가능.
"""
import os
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from data.models import WordEntry


_HEADERS = [
    "날짜", "유형", "표현", "뜻", "예문",
    "예문_번역", "레이블", "복습필요", "복습_횟수", "다음복습일", "기억강도",
]


class WordbookRepository:
    def __init__(self, file_path: str | Path):
        self.path = Path(file_path)
        self._ensure_file()

    # ── 읽기 ──────────────────────────────────────────────────────────────
    def load_all(self) -> list[WordEntry]:
        wb = openpyxl.load_workbook(self.path)
        ws = wb.active
        entries = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            entry = self._row_to_entry(row, idx)
            if entry:
                entries.append(entry)
        wb.close()
        return entries

    # ── 쓰기 ──────────────────────────────────────────────────────────────
    def add(self, entry: WordEntry) -> WordEntry:
        wb = openpyxl.load_workbook(self.path)
        ws = wb.active
        next_row = ws.max_row + 1
        ws.append(self._entry_to_row(entry))
        wb.save(self.path)
        wb.close()
        entry.row_index = next_row
        return entry

    def update(self, entry: WordEntry) -> None:
        """row_index가 있는 기존 항목 업데이트."""
        if entry.row_index is None:
            raise ValueError("row_index가 없는 항목은 update할 수 없습니다.")
        wb = openpyxl.load_workbook(self.path)
        ws = wb.active
        row_data = self._entry_to_row(entry)
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=entry.row_index, column=col, value=value)
        wb.save(self.path)
        wb.close()

    def delete(self, entry: WordEntry) -> None:
        if entry.row_index is None:
            raise ValueError("row_index가 없는 항목은 delete할 수 없습니다.")
        wb = openpyxl.load_workbook(self.path)
        ws = wb.active
        ws.delete_rows(entry.row_index)
        wb.save(self.path)
        wb.close()

    # ── 내부 변환 ──────────────────────────────────────────────────────────
    def _row_to_entry(self, row: tuple, row_index: int) -> Optional[WordEntry]:
        if not row or row[2] is None:   # 표현이 없으면 건너뜀
            return None
        try:
            return WordEntry(
                added_date=_to_date(row[0]),
                word_type=str(row[1] or "vocabulary"),
                word=str(row[2]),
                meaning=str(row[3] or ""),
                example=str(row[4] or ""),
                example_translation=str(row[5] or ""),
                label=str(row[6] or "초기"),
                needs_review=(str(row[7] or "N").upper() == "Y"),
                review_count=int(row[8] or 0),
                next_review=_to_date(row[9]),
                memory_strength=int(row[10] or 0),
                row_index=row_index,
            )
        except Exception:
            return None

    def _entry_to_row(self, entry: WordEntry) -> list:
        return [
            entry.added_date.isoformat(),
            entry.word_type,
            entry.word,
            entry.meaning,
            entry.example,
            entry.example_translation,
            entry.label,
            "Y" if entry.needs_review else "N",
            entry.review_count,
            entry.next_review.isoformat(),
            entry.memory_strength,
        ]

    def _ensure_file(self) -> None:
        if self.path.exists():
            return
        wb = Workbook()
        ws = wb.active
        ws.append(_HEADERS)
        wb.save(self.path)
        wb.close()


def _to_date(value) -> date:
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        return date.fromisoformat(value.strip())
    return date.today()
